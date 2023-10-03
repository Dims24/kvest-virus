import os
import pandas as pd
import psycopg2
from datetime import date

from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter


def calculate_points(row):
    # Подсчет баллов на основе полей question
    points = 0
    for i in range(1, 10):
        if row[f'Задание{i}'] == "выполнено":
            points += 1
    return points


def map_boolean_to_status(value):
    if value == True:
        return 'выполнено'
    else:
        return 'не выполнено'


def export_data():
    conn = psycopg2.connect(
        dbname='virus',
        user='virus',
        password='virus',
        host='localhost',
        port='5010'
    )
    cur = conn.cursor()

    # Выбор необходимых полей из таблицы tasks
    query = '''
    SELECT 
        chat_id,
        name,
        question_1,
        question_2,
        question_3,
        question_4,
        question_5,
        question_6,
        question_7,
        question_8,
        question_9
    FROM tasks
    '''
    cur.execute(query)
    data = cur.fetchall()

    # Создание DataFrame
    columns = [
        'Chat id',
        'Название команды',
        'Задание1',
        'Задание2',
        'Задание3',
        'Задание4',
        'Задание5',
        'Задание6',
        'Задание7',
        'Задание8',
        'Задание9',
    ]

    df = pd.DataFrame(data, columns=columns)

    # Преобразование значений "true" и "false" в "выполнено" и "не выполнено"
    df[['Задание1', 'Задание2', 'Задание3', 'Задание4', 'Задание5', 'Задание6', 'Задание7', 'Задание8', 'Задание9']] = \
    df[['Задание1', 'Задание2', 'Задание3', 'Задание4', 'Задание5', 'Задание6', 'Задание7', 'Задание8',
        'Задание9']].applymap(map_boolean_to_status)

    # Вычисление баллов с использованием функции calculate_points
    df['Баллы'] = df.apply(calculate_points, axis=1)

    current_date = date.today()

    # Экспорт DataFrame в Excel
    df.to_excel(f'{current_date}.xlsx', index=False)
    file_path = os.path.join(os.getcwd(), f'{current_date}.xlsx')

    # Экспорт DataFrame в Excel
    writer = pd.ExcelWriter(file_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='Sheet1')

    # Получение объекта рабочей книги
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']

    # Создание объекта для заливки ячеек зеленым цветом
    fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

    # Поиск всех ячеек со значением "выполнено" и применение заливки
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=3, max_col=12):
        for cell in row:
            if cell.value == 'выполнено':
                cell.fill = fill
        # Автоподбор ширины столбцов
        for col in worksheet.columns:
            max_length = 0
            column = get_column_letter(col[0].column)
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            worksheet.column_dimensions[column].width = adjusted_width
    # Сохранение файла Excel
    writer.save()
    writer.close()

    return file_path
    # return path
